#!/usr/bin/env python3
"""
Backward Mapping Script

This script merges old and new databases using intelligent path matching.
It handles conflicts (1-to-many, many-to-many) and creates a highlighted Excel output.
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Merge old and new databases with intelligent path matching.'
    )
    parser.add_argument(
        '--new-excel',
        type=str,
        help='Path to the new database Excel file'
    )
    parser.add_argument(
        '--old-excel',
        type=str,
        help='Path to the old database Excel file'
    )
    return parser.parse_args()


def get_excel_paths() -> Tuple[Path, Path]:
    """
    Get paths to both Excel files from arguments or user input.

    Returns:
        Tuple of (new_excel_path, old_excel_path)
    """
    args = parse_arguments()

    # Default paths for interactive mode
    default_new_path = Path(r"C:\Users\LEGION\Documents\GIT\FICZall-Find-Index-Clean-Zip-all-Dicom-Files-and-Reports\data\consolidated_summaries\consolidated_patient_summary_20251014_AbCT-v1.xlsx")
    default_old_path = Path(r"C:\Users\LEGION\Documents\GIT\FICZall-Find-Index-Clean-Zip-all-Dicom-Files-and-Reports\data\consolidated_summaries\PanCanAID Master 2024.xlsx")

    # Get new Excel path
    if args.new_excel:
        new_excel = Path(args.new_excel)
    else:
        print("\n" + "=" * 60)
        print("NEW DATABASE")
        print("=" * 60)
        print(f"Default: {default_new_path}")
        new_input = input("Enter path (or press Enter for default): ").strip()

        if new_input:
            new_excel = Path(new_input)
        else:
            new_excel = default_new_path
            print(f"Using default: {new_excel}")

    if not new_excel.exists():
        print(f"Error: New Excel file not found: {new_excel}")
        sys.exit(1)

    # Get old Excel path
    if args.old_excel:
        old_excel = Path(args.old_excel)
    else:
        print("\n" + "=" * 60)
        print("OLD DATABASE")
        print("=" * 60)
        print(f"Default: {default_old_path}")
        old_input = input("Enter path (or press Enter for default): ").strip()

        if old_input:
            old_excel = Path(old_input)
        else:
            old_excel = default_old_path
            print(f"Using default: {old_excel}")

    if not old_excel.exists():
        print(f"Error: Old Excel file not found: {old_excel}")
        sys.exit(1)

    return new_excel, old_excel


def load_dataframes(new_excel: Path, old_excel: Path):
    """
    Load both Excel files into pandas DataFrames.

    Args:
        new_excel: Path to new database Excel
        old_excel: Path to old database Excel

    Returns:
        Tuple of (new_df, old_df)
    """
    try:
        import pandas as pd
    except ImportError:
        print("Error: pandas is required. Install it: pip install pandas openpyxl")
        sys.exit(1)

    try:
        new_df = pd.read_excel(new_excel)
        print(f"Loaded new database: {len(new_df)} rows")
    except Exception as e:
        print(f"Error reading new Excel file: {e}")
        sys.exit(1)

    try:
        old_df = pd.read_excel(old_excel)
        print(f"Loaded old database: {len(old_df)} rows")
    except Exception as e:
        print(f"Error reading old Excel file: {e}")
        sys.exit(1)

    return new_df, old_df


def validate_required_columns(new_df, old_df):
    """
    Validate that required columns exist in both DataFrames.

    Args:
        new_df: New database DataFrame
        old_df: Old database DataFrame

    Raises:
        SystemExit: If critical required columns are missing
    """
    # Critical required columns in new database
    new_required = ['Hospital', 'dicom_folder_path']
    missing_new = [col for col in new_required if col not in new_df.columns]
    if missing_new:
        print(f"Error: New database missing critical columns: {', '.join(missing_new)}")
        print(f"Available columns: {', '.join(new_df.columns)}")
        sys.exit(1)

    # Critical required columns in old database
    old_required = ['Hospital', 'CT_No_Final']
    missing_old = [col for col in old_required if col not in old_df.columns]
    if missing_old:
        print(f"Error: Old database missing critical columns: {', '.join(missing_old)}")
        print(f"Available columns: {', '.join(old_df.columns)}")
        sys.exit(1)

    # Check optional columns and warn if missing
    optional_new = ['Unique_number', 'is_duplicate', 'is_duplicate_first_occurrence_un', 'study_date']
    optional_old = ['dir_to_root', 'DONOTPUBLISH_ID_fromCT', 'ID_HOSPITAL', 'date']

    missing_optional_new = [col for col in optional_new if col not in new_df.columns]
    missing_optional_old = [col for col in optional_old if col not in old_df.columns]

    if missing_optional_new:
        print(f"Note: Optional columns missing in new database: {', '.join(missing_optional_new)}")
        print("  (Some features may be skipped)")

    if missing_optional_old:
        print(f"Note: Optional columns missing in old database: {', '.join(missing_optional_old)}")
        print("  (Some matching strategies may be skipped)")

    print("All critical columns present.")


def normalize_path(path: str) -> str:
    """
    Normalize a path by removing drive letter and converting to standard format.

    Args:
        path: Original path string

    Returns:
        Normalized path without drive letter, using forward slashes, lowercased
    """
    if not path or not isinstance(path, str):
        return ""

    # Remove drive letter (e.g., E:, F:, D:)
    path_normalized = re.sub(r'^[A-Za-z]:', '', path.strip())

    # Convert backslashes to forward slashes
    path_normalized = path_normalized.replace('\\', '/')

    # Remove leading/trailing slashes
    path_normalized = path_normalized.strip('/')

    # Convert to lowercase for comparison
    path_normalized = path_normalized.lower()

    return path_normalized


def parse_patient_id(id_hospital: str) -> str:
    """
    Parse patient ID from ID_HOSPITAL by removing last underscore and everything after.

    Args:
        id_hospital: ID_HOSPITAL value (e.g., "12973355_EmamKh")

    Returns:
        Parsed patient ID (e.g., "12973355")
    """
    if not id_hospital or not isinstance(id_hospital, str):
        return ""

    id_str = str(id_hospital).strip()

    # Find last underscore and remove everything after it
    if '_' in id_str:
        return id_str.rsplit('_', 1)[0]

    return id_str


def dates_match(date1: str, date2: str) -> bool:
    """
    Check if two dates match after normalization.

    Args:
        date1: First date string
        date2: Second date string

    Returns:
        True if dates match, False otherwise
    """
    import pandas as pd

    if pd.isna(date1) or pd.isna(date2):
        return False

    try:
        # Convert to strings and normalize
        date1_str = str(date1).strip()
        date2_str = str(date2).strip()

        if not date1_str or not date2_str:
            return False

        # Remove common separators and compare
        date1_clean = re.sub(r'[-/\s]', '', date1_str)
        date2_clean = re.sub(r'[-/\s]', '', date2_str)

        return date1_clean == date2_clean
    except:
        return False


def is_path_match(new_path: str, old_path: str) -> bool:
    """
    Check if two paths match using intelligent logic.

    The new path can be:
    1. Exactly the same as old path (after normalization)
    2. A subdirectory of the old path

    Args:
        new_path: Path from new database (dicom_folder_path)
        old_path: Path from old database (dir_to_root)

    Returns:
        True if paths match according to the logic
    """
    new_norm = normalize_path(new_path)
    old_norm = normalize_path(old_path)

    if not new_norm or not old_norm:
        return False

    # Case 1: Exact match
    if new_norm == old_norm:
        return True

    # Case 2: New path is a subdirectory of old path
    # The new path should start with the old path
    if new_norm.startswith(old_norm + '/'):
        return True

    return False


def find_matches(new_df, old_df) -> Dict[int, List[int]]:
    """
    Find matching rows between new and old databases using multiple strategies.

    Strategies:
    1. Path-based matching (primary) + date validation
    2. ID-based matching (for rows without dir_to_root) + date validation
       a. DONOTPUBLISH_ID_fromCT matching
       b. Parsed ID_HOSPITAL matching

    Date validation: If both study_date (new) and date (old) exist, they must match.

    Args:
        new_df: New database DataFrame
        old_df: Old database DataFrame

    Returns:
        Dictionary mapping new_row_index -> list of old_row_indices
    """
    import pandas as pd

    matches = defaultdict(list)
    total_new = len(new_df)

    # Check if date columns exist
    has_new_date = 'study_date' in new_df.columns
    has_old_date = 'date' in old_df.columns
    date_validation_enabled = has_new_date and has_old_date

    if date_validation_enabled:
        print("\nDate validation enabled: Matches must have matching dates (if both exist)")
    else:
        print("\nDate validation disabled: Missing date columns")

    print("\nFinding matches between databases...")
    print("Strategy 1: Path-based matching...")

    # Strategy 1: Path-based matching
    for new_idx, new_row in new_df.iterrows():
        if (new_idx + 1) % 100 == 0:
            print(f"  Processing new database row {new_idx + 1}/{total_new}...", end='\r')

        new_hospital = new_row.get('Hospital', '')
        new_path = new_row.get('dicom_folder_path', '')
        new_date = new_row.get('study_date', '') if has_new_date else None

        # Skip if missing hospital
        if pd.isna(new_hospital):
            continue

        # Find matching rows in old database
        for old_idx, old_row in old_df.iterrows():
            old_hospital = old_row.get('Hospital', '')
            old_path = old_row.get('dir_to_root', '')
            old_date = old_row.get('date', '') if has_old_date else None

            # Check hospital match first
            if pd.isna(old_hospital) or str(new_hospital).strip() != str(old_hospital).strip():
                continue

            # Try path matching if both paths exist
            if not pd.isna(new_path) and not pd.isna(old_path) and old_path:
                if is_path_match(str(new_path), str(old_path)):
                    # Apply date validation only if both dates exist
                    if date_validation_enabled and not pd.isna(new_date) and not pd.isna(old_date):
                        if not dates_match(new_date, old_date):
                            continue  # Skip this match if dates don't match

                    if old_idx not in matches[new_idx]:
                        matches[new_idx].append(old_idx)

    print(f"  Path-based matching completed.                          ")

    # Strategy 2: ID-based matching for old rows without dir_to_root
    print("Strategy 2: ID-based matching (for rows without dir_to_root)...")

    # Get columns for ID matching in new database
    new_id_cols = []
    if 'ID_fromCT' in new_df.columns:
        new_id_cols.append('ID_fromCT')
    if 'DONOTPUBLISH_ID_fromCT' in new_df.columns:
        new_id_cols.append('DONOTPUBLISH_ID_fromCT')
    if 'patient_id' in new_df.columns:
        new_id_cols.append('patient_id')

    for old_idx, old_row in old_df.iterrows():
        old_hospital = old_row.get('Hospital', '')
        old_path = old_row.get('dir_to_root', '')
        old_date = old_row.get('date', '') if has_old_date else None

        # Skip if hospital is missing or if this row already has path matches
        if pd.isna(old_hospital):
            continue

        # Only use ID matching if dir_to_root is empty
        if not pd.isna(old_path) and str(old_path).strip():
            continue

        # Try matching with DONOTPUBLISH_ID_fromCT
        old_id_ct = old_row.get('DONOTPUBLISH_ID_fromCT', '')
        if not pd.isna(old_id_ct) and str(old_id_ct).strip():
            old_id_ct_str = str(old_id_ct).strip()

            for new_idx, new_row in new_df.iterrows():
                new_hospital = new_row.get('Hospital', '')
                new_date = new_row.get('study_date', '') if has_new_date else None

                # Check hospital match
                if pd.isna(new_hospital) or str(new_hospital).strip() != str(old_hospital).strip():
                    continue

                # Skip if already matched
                if old_idx in matches.get(new_idx, []):
                    continue

                # Try matching with any ID column in new database
                for id_col in new_id_cols:
                    new_id = new_row.get(id_col, '')
                    if not pd.isna(new_id) and str(new_id).strip() == old_id_ct_str:
                        # Apply date validation only if both dates exist
                        if date_validation_enabled and not pd.isna(new_date) and not pd.isna(old_date):
                            if not dates_match(new_date, old_date):
                                break  # Skip this match if dates don't match

                        if old_idx not in matches[new_idx]:
                            matches[new_idx].append(old_idx)
                        break

        # If still not matched, try ID_HOSPITAL
        elif 'ID_HOSPITAL' in old_df.columns:
            old_id_hospital = old_row.get('ID_HOSPITAL', '')
            if not pd.isna(old_id_hospital) and str(old_id_hospital).strip():
                parsed_id = parse_patient_id(str(old_id_hospital))

                if parsed_id:
                    for new_idx, new_row in new_df.iterrows():
                        new_hospital = new_row.get('Hospital', '')
                        new_date = new_row.get('study_date', '') if has_new_date else None

                        # Check hospital match
                        if pd.isna(new_hospital) or str(new_hospital).strip() != str(old_hospital).strip():
                            continue

                        # Skip if already matched
                        if old_idx in matches.get(new_idx, []):
                            continue

                        # Try matching with any ID column in new database
                        for id_col in new_id_cols:
                            new_id = new_row.get(id_col, '')
                            if not pd.isna(new_id) and str(new_id).strip() == parsed_id:
                                # Apply date validation if both dates exist
                                if date_validation_enabled and not pd.isna(new_date) and not pd.isna(old_date):
                                    if not dates_match(new_date, old_date):
                                        break  # Skip this match if dates don't match

                                if old_idx not in matches[new_idx]:
                                    matches[new_idx].append(old_idx)
                                break

    print(f"  ID-based matching completed.                            ")
    print(f"  Total matches found: {len(matches)} new rows matched")
    return dict(matches)


def build_duplicate_groups(new_df) -> Dict[str, List[int]]:
    """
    Build duplicate groups from the new database.

    Args:
        new_df: New database DataFrame

    Returns:
        Dictionary mapping unique_number -> list of indices (first_occurrence + duplicates)
    """
    import pandas as pd

    duplicate_groups = defaultdict(list)

    # Check if required columns exist
    if 'is_duplicate' not in new_df.columns:
        return dict(duplicate_groups)

    for idx, row in new_df.iterrows():
        is_dup = row.get('is_duplicate', '')

        if pd.isna(is_dup):
            continue

        is_dup_str = str(is_dup).strip().lower()

        if is_dup_str == 'first_occurrence':
            # This is the first occurrence
            unique_num = row.get('Unique_number', '') if 'Unique_number' in new_df.columns else ''
            if not pd.isna(unique_num):
                duplicate_groups[str(unique_num)].append(idx)

        elif is_dup_str == 'duplicate':
            # This is a duplicate - find its first occurrence
            first_occ_un = row.get('is_duplicate_first_occurrence_un', '') if 'is_duplicate_first_occurrence_un' in new_df.columns else ''
            if not pd.isna(first_occ_un):
                duplicate_groups[str(first_occ_un)].append(idx)

    return dict(duplicate_groups)


def propagate_duplicate_matches(matches: Dict[int, List[int]], new_df) -> Tuple[Dict[int, List[int]], List[int]]:
    """
    Propagate matches to duplicate pairs.

    If one row in a duplicate group has a match but others don't, copy the match.

    Args:
        matches: Current matches dictionary
        new_df: New database DataFrame

    Returns:
        Tuple of (updated_matches, list of indices that got propagated matches - cyan)
    """
    duplicate_groups = build_duplicate_groups(new_df)
    propagated_indices = []

    if not duplicate_groups:
        print("  No duplicate groups found.")
        return matches, propagated_indices

    print(f"  Found {len(duplicate_groups)} duplicate groups")

    for unique_num, group_indices in duplicate_groups.items():
        if len(group_indices) < 2:
            continue

        # Find which rows in the group have matches
        matched_in_group = [idx for idx in group_indices if idx in matches and matches[idx]]
        unmatched_in_group = [idx for idx in group_indices if idx not in matches or not matches[idx]]

        # If some have matches and others don't, propagate
        if matched_in_group and unmatched_in_group:
            # Use the first matched row's matches as the source
            source_matches = matches[matched_in_group[0]]

            for unmatch_idx in unmatched_in_group:
                matches[unmatch_idx] = source_matches.copy()
                propagated_indices.append(unmatch_idx)

    print(f"  Propagated matches to {len(propagated_indices)} duplicate rows")
    return matches, propagated_indices


def validate_dates(matches: Dict[int, List[int]], new_df, old_df) -> List[int]:
    """
    Validate dates for matched rows.

    Check if study_date in new database matches date in old database.

    Args:
        matches: Matches dictionary
        new_df: New database DataFrame
        old_df: Old database DataFrame

    Returns:
        List of new row indices with date mismatches (dark green highlighting)
    """
    import pandas as pd

    date_mismatch_indices = []

    # Check if required columns exist
    if 'study_date' not in new_df.columns or 'date' not in old_df.columns:
        print("  Date validation skipped (required columns not found)")
        return date_mismatch_indices

    print("  Validating dates for matched rows...")

    for new_idx, old_indices in matches.items():
        if not old_indices:
            continue

        new_date = new_df.loc[new_idx].get('study_date', '')

        if pd.isna(new_date):
            continue

        # Check against all matched old rows
        for old_idx in old_indices:
            old_date = old_df.loc[old_idx].get('date', '')

            if pd.isna(old_date) or not str(old_date).strip():
                continue

            # Convert dates to comparable format
            new_date_str = str(new_date).strip()
            old_date_str = str(old_date).strip()

            # Try to normalize dates for comparison (handle different formats)
            try:
                # Remove common separators and compare
                new_date_clean = re.sub(r'[-/\s]', '', new_date_str)
                old_date_clean = re.sub(r'[-/\s]', '', old_date_str)

                if new_date_clean != old_date_clean:
                    if new_idx not in date_mismatch_indices:
                        date_mismatch_indices.append(new_idx)
            except:
                pass

    print(f"  Found {len(date_mismatch_indices)} rows with date mismatches")
    return date_mismatch_indices


def analyze_conflicts(matches: Dict[int, List[int]], old_df) -> Dict:
    """
    Analyze matches to detect conflicts.

    Args:
        matches: Dictionary mapping new_row_index -> list of old_row_indices
        old_df: Old database DataFrame

    Returns:
        Dictionary containing conflict analysis
    """
    # Reverse mapping: old_idx -> list of new_idx
    reverse_matches = defaultdict(list)
    for new_idx, old_indices in matches.items():
        for old_idx in old_indices:
            reverse_matches[old_idx].append(new_idx)

    conflict_info = {
        'one_to_one': [],        # Perfect matches
        'one_to_many': [],       # 1 old -> many new (yellow)
        'many_to_one': [],       # Many old -> 1 new (red)
        'many_to_many': [],      # Many old -> many new (red)
        'unmatched_new': [],     # New rows with no match
        'unmatched_old': []      # Old rows with no match
    }

    # Analyze forward matches (new -> old)
    for new_idx, old_indices in matches.items():
        if len(old_indices) == 1:
            old_idx = old_indices[0]
            # Check if it's truly 1-to-1
            if len(reverse_matches[old_idx]) == 1:
                conflict_info['one_to_one'].append((new_idx, old_idx))

    # Analyze reverse matches (old -> new)
    for old_idx, new_indices in reverse_matches.items():
        if len(new_indices) == 1:
            # Already handled in one_to_one
            pass
        else:
            # 1 old -> many new (yellow - acceptable)
            conflict_info['one_to_many'].append((old_idx, new_indices))

    # Find many-to-X patterns (red - problematic)
    processed_new = set()
    for new_idx, old_indices in matches.items():
        if new_idx in processed_new:
            continue

        if len(old_indices) > 1:
            # Many old -> ? new
            # Check if any of these old indices map to other new indices
            all_new_for_these_old = set()
            for old_idx in old_indices:
                all_new_for_these_old.update(reverse_matches[old_idx])

            if len(all_new_for_these_old) == 1:
                # Many old -> 1 new (red)
                conflict_info['many_to_one'].append((list(old_indices), new_idx))
            else:
                # Many old -> many new (red)
                conflict_info['many_to_many'].append((list(old_indices), list(all_new_for_these_old)))

            processed_new.update(all_new_for_these_old)

    return conflict_info


def create_merged_dataframe(new_df, old_df, matches: Dict[int, List[int]],
                           conflict_info: Dict, propagated_indices: List[int],
                           date_mismatch_indices: List[int]):
    """
    Create merged DataFrame with outer join style.

    Args:
        new_df: New database DataFrame
        old_df: Old database DataFrame
        matches: Dictionary mapping new_row_index -> list of old_row_indices
        conflict_info: Conflict analysis results
        propagated_indices: List of new indices that got propagated matches (cyan)
        date_mismatch_indices: List of new indices with date mismatches (dark green)

    Returns:
        Tuple of (merged_df, highlight_info)
    """
    import pandas as pd

    # Prepare columns to merge from old database
    # Only include columns that actually exist in old_df
    potential_columns = {
        'dir_to_root': 'backward_dir_to_root',
        'DONOTPUBLISH_ID_fromCT': 'backward_DONOTPUBLISH_ID_fromCT',
        'CT_No_Final': 'backward_CT_No_Final',
        'ID_HOSPITAL': 'backward_ID_HOSPITAL',
        'date': 'backward_date'
    }

    # Filter to only columns that exist in old database
    rename_mapping = {
        old_col: new_col
        for old_col, new_col in potential_columns.items()
        if old_col in old_df.columns
    }

    # Initialize backward columns in new_df
    for old_col, new_col in rename_mapping.items():
        new_df[new_col] = None

    merged_rows = []
    highlight_info = []  # List of (row_index, color)
    new_idx_to_merged_idx = {}  # Track mapping from new_idx to merged row index

    # Helper function to determine color priority
    def get_color_for_row(new_idx):
        """Determine highlighting color with priority: red > yellow > cyan > dark_green"""
        # Check if it's in any conflict category
        for _, idx in conflict_info['many_to_one']:
            if idx == new_idx:
                return 'red'
        for _, indices in conflict_info['many_to_many']:
            if new_idx in indices:
                return 'red'
        for _, indices in conflict_info['one_to_many']:
            if new_idx in indices:
                return 'yellow'
        if new_idx in propagated_indices:
            return 'cyan'
        if new_idx in date_mismatch_indices:
            return 'dark_green'
        return None

    # Process 1-to-1 matches
    print("\nProcessing 1-to-1 matches...")
    for new_idx, old_idx in conflict_info['one_to_one']:
        row = new_df.loc[new_idx].copy()
        for old_col, new_col in rename_mapping.items():
            row[new_col] = old_df.loc[old_idx, old_col]
        merged_rows.append(row)
        color = get_color_for_row(new_idx)
        highlight_info.append((len(merged_rows) - 1, color))
        new_idx_to_merged_idx[new_idx] = len(merged_rows) - 1

    # Process 1-to-many matches (yellow)
    print("Processing 1-to-many matches (yellow)...")
    for old_idx, new_indices in conflict_info['one_to_many']:
        for new_idx in new_indices:
            row = new_df.loc[new_idx].copy()
            for old_col, new_col in rename_mapping.items():
                row[new_col] = old_df.loc[old_idx, old_col]
            merged_rows.append(row)
            color = get_color_for_row(new_idx)
            highlight_info.append((len(merged_rows) - 1, color))
            new_idx_to_merged_idx[new_idx] = len(merged_rows) - 1

    # Process many-to-one matches (red - don't merge, keep separate)
    print("Processing many-to-one matches (red)...")
    for old_indices, new_idx in conflict_info['many_to_one']:
        # Add new row without merge
        row = new_df.loc[new_idx].copy()
        merged_rows.append(row)
        color = get_color_for_row(new_idx)
        highlight_info.append((len(merged_rows) - 1, color))
        new_idx_to_merged_idx[new_idx] = len(merged_rows) - 1

    # Process many-to-many matches (red - don't merge, keep separate)
    print("Processing many-to-many matches (red)...")
    for old_indices, new_indices in conflict_info['many_to_many']:
        for new_idx in new_indices:
            row = new_df.loc[new_idx].copy()
            merged_rows.append(row)
            color = get_color_for_row(new_idx)
            highlight_info.append((len(merged_rows) - 1, color))
            new_idx_to_merged_idx[new_idx] = len(merged_rows) - 1

    # Add unmatched new rows
    print("Adding unmatched new database rows...")
    matched_new = set()
    for new_idx, _ in conflict_info['one_to_one']:
        matched_new.add(new_idx)
    for _, new_indices in conflict_info['one_to_many']:
        matched_new.update(new_indices)
    for _, new_idx in conflict_info['many_to_one']:
        matched_new.add(new_idx)
    for _, new_indices in conflict_info['many_to_many']:
        matched_new.update(new_indices)

    for new_idx in new_df.index:
        if new_idx not in matched_new:
            row = new_df.loc[new_idx].copy()
            merged_rows.append(row)
            color = get_color_for_row(new_idx)
            highlight_info.append((len(merged_rows) - 1, color))
            new_idx_to_merged_idx[new_idx] = len(merged_rows) - 1

    # Add unmatched old rows
    print("Adding unmatched old database rows...")
    matched_old = set()
    for _, old_idx in conflict_info['one_to_one']:
        matched_old.add(old_idx)
    for old_idx, _ in conflict_info['one_to_many']:
        matched_old.add(old_idx)
    for old_indices, _ in conflict_info['many_to_one']:
        matched_old.update(old_indices)
    for old_indices, _ in conflict_info['many_to_many']:
        matched_old.update(old_indices)

    for old_idx in old_df.index:
        if old_idx not in matched_old:
            # Create row with old data only
            row = pd.Series({col: None for col in new_df.columns})
            for old_col, new_col in rename_mapping.items():
                if old_col in old_df.columns:
                    row[new_col] = old_df.loc[old_idx, old_col]
            # Add Hospital from old database
            if 'Hospital' in old_df.columns:
                row['Hospital'] = old_df.loc[old_idx, 'Hospital']
            merged_rows.append(row)
            highlight_info.append((len(merged_rows) - 1, None))

    # Create merged DataFrame
    merged_df = pd.DataFrame(merged_rows)
    merged_df.reset_index(drop=True, inplace=True)

    return merged_df, highlight_info


def save_conflict_log(conflict_info: Dict, new_df, old_df, output_path: Path):
    """
    Save conflict information to a JSON file.

    Args:
        conflict_info: Conflict analysis results
        new_df: New database DataFrame
        old_df: Old database DataFrame
        output_path: Path where JSON should be saved
    """
    log_data = {
        'summary': {
            'one_to_one_matches': len(conflict_info['one_to_one']),
            'one_to_many_matches': len(conflict_info['one_to_many']),
            'many_to_one_matches': len(conflict_info['many_to_one']),
            'many_to_many_matches': len(conflict_info['many_to_many'])
        },
        'conflicts': []
    }

    # Log 1-to-many conflicts (yellow)
    for old_idx, new_indices in conflict_info['one_to_many']:
        conflict_entry = {
            'type': '1-to-many (yellow)',
            'old_row': int(old_idx),
            'old_CT_No_Final': str(old_df.loc[old_idx, 'CT_No_Final']) if 'CT_No_Final' in old_df.columns else None,
            'new_rows': [int(idx) for idx in new_indices],
            'new_Unique_numbers': []
        }
        if 'Unique_number' in new_df.columns:
            conflict_entry['new_Unique_numbers'] = [
                str(new_df.loc[idx, 'Unique_number']) for idx in new_indices
            ]
        log_data['conflicts'].append(conflict_entry)

    # Log many-to-one conflicts (red)
    for old_indices, new_idx in conflict_info['many_to_one']:
        conflict_entry = {
            'type': 'many-to-1 (red - not merged)',
            'old_rows': [int(idx) for idx in old_indices],
            'old_CT_No_Finals': [],
            'new_row': int(new_idx),
            'new_Unique_number': None
        }
        if 'CT_No_Final' in old_df.columns:
            conflict_entry['old_CT_No_Finals'] = [
                str(old_df.loc[idx, 'CT_No_Final']) for idx in old_indices
            ]
        if 'Unique_number' in new_df.columns:
            conflict_entry['new_Unique_number'] = str(new_df.loc[new_idx, 'Unique_number'])
        log_data['conflicts'].append(conflict_entry)

    # Log many-to-many conflicts (red)
    for old_indices, new_indices in conflict_info['many_to_many']:
        conflict_entry = {
            'type': 'many-to-many (red - not merged)',
            'old_rows': [int(idx) for idx in old_indices],
            'old_CT_No_Finals': [],
            'new_rows': [int(idx) for idx in new_indices],
            'new_Unique_numbers': []
        }
        if 'CT_No_Final' in old_df.columns:
            conflict_entry['old_CT_No_Finals'] = [
                str(old_df.loc[idx, 'CT_No_Final']) for idx in old_indices
            ]
        if 'Unique_number' in new_df.columns:
            conflict_entry['new_Unique_numbers'] = [
                str(new_df.loc[idx, 'Unique_number']) for idx in new_indices
            ]
        log_data['conflicts'].append(conflict_entry)

    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(log_data, f, indent=2, ensure_ascii=False)
        print(f"\nConflict log saved to: {output_path}")
    except Exception as e:
        print(f"Warning: Could not save conflict log: {e}")


def save_merged_excel(merged_df, highlight_info: List[Tuple[int, Optional[str]]],
                     output_path: Path):
    """
    Save merged DataFrame to Excel with highlighting.

    Color scheme:
    - Red: Many-to-one or many-to-many conflicts (not merged)
    - Yellow: One-to-many matches (acceptable duplicates)
    - Cyan: Propagated duplicate matches
    - Dark Green: Date mismatches in matched rows

    Args:
        merged_df: Merged DataFrame
        highlight_info: List of (row_index, color) tuples
        output_path: Path where Excel should be saved
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        print("Warning: openpyxl not available, saving without highlighting")
        merged_df.to_excel(output_path, index=False)
        print(f"\nMerged database saved to: {output_path}")
        return

    # Save to Excel first
    merged_df.to_excel(output_path, index=False, engine='openpyxl')

    # Load workbook for highlighting
    wb = load_workbook(output_path)
    ws = wb.active

    # Create fill styles
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')

    # Apply highlighting
    color_counts = {
        'yellow': 0,
        'red': 0,
        'cyan': 0,
        'dark_green': 0
    }

    for row_idx, color in highlight_info:
        if color is None:
            continue

        excel_row = row_idx + 2  # +1 for 0-index, +1 for header

        fill_style = None
        if color == 'yellow':
            fill_style = yellow_fill
        elif color == 'red':
            fill_style = red_fill
        elif color == 'cyan':
            fill_style = cyan_fill
        elif color == 'dark_green':
            fill_style = dark_green_fill

        if fill_style:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = fill_style
            color_counts[color] = color_counts.get(color, 0) + 1

    wb.save(output_path)
    print(f"\nMerged database saved to: {output_path}")
    print(f"  - Red highlighted rows (conflicts): {color_counts['red']}")
    print(f"  - Yellow highlighted rows (1-to-many): {color_counts['yellow']}")
    print(f"  - Cyan highlighted rows (propagated duplicates): {color_counts['cyan']}")
    print(f"  - Dark green highlighted rows (date mismatches): {color_counts['dark_green']}")


def print_summary(conflict_info: Dict, merged_df):
    """Print summary of the mapping process."""
    print("\n" + "=" * 60)
    print("MAPPING SUMMARY")
    print("=" * 60)
    print(f"1-to-1 matches (clean): {len(conflict_info['one_to_one'])}")
    print(f"1-to-many matches (yellow): {len(conflict_info['one_to_many'])}")
    print(f"Many-to-1 matches (red, not merged): {len(conflict_info['many_to_one'])}")
    print(f"Many-to-many matches (red, not merged): {len(conflict_info['many_to_many'])}")
    print(f"\nTotal rows in merged database: {len(merged_df)}")
    print("=" * 60)


def main():
    """Main function to orchestrate the backward mapping process."""
    print("=" * 60)
    print("Backward Mapping Script")
    print("=" * 60)

    # Get Excel paths
    new_excel, old_excel = get_excel_paths()
    print(f"\nNew database: {new_excel}")
    print(f"Old database: {old_excel}")

    # Load DataFrames
    new_df, old_df = load_dataframes(new_excel, old_excel)

    # Validate columns
    validate_required_columns(new_df, old_df)

    # Find matches (includes path-based and ID-based matching)
    matches = find_matches(new_df, old_df)
    print(f"\nFound {len(matches)} new database rows with matches")

    # Propagate matches to duplicate pairs
    print("\nHandling duplicate propagation...")
    matches, propagated_indices = propagate_duplicate_matches(matches, new_df)

    # Analyze conflicts
    print("\nAnalyzing conflicts...")
    conflict_info = analyze_conflicts(matches, old_df)

    # Validate dates for matched rows
    print("\nValidating dates...")
    date_mismatch_indices = validate_dates(matches, new_df, old_df)

    # Create merged DataFrame
    print("\nCreating merged database...")
    merged_df, highlight_info = create_merged_dataframe(
        new_df.copy(), old_df.copy(), matches, conflict_info,
        propagated_indices, date_mismatch_indices
    )

    # Generate output paths
    output_excel = new_excel.parent / f"{new_excel.stem}_Map2Old{new_excel.suffix}"
    output_json = new_excel.parent / f"{new_excel.stem}_Map2Old_conflicts.json"

    # Save conflict log
    save_conflict_log(conflict_info, new_df, old_df, output_json)

    # Save merged Excel
    save_merged_excel(merged_df, highlight_info, output_excel)

    # Print summary
    print_summary(conflict_info, merged_df)

    print("\n" + "=" * 60)
    print("HIGHLIGHTING LEGEND:")
    print("  - Red: Conflicts (many-to-1 or many-to-many) - NOT merged")
    print("  - Yellow: 1-to-many matches - merged successfully")
    print("  - Cyan: Duplicate pairs with propagated matches")
    print("  - Dark Green: Date mismatches between databases")
    print("=" * 60)
    print("\nProcess completed successfully!")
    print("=" * 60)


if __name__ == "__main__":
    main()
